'''
一行に最大 4 組の座標が登録されているワークシートのデータを
一行に 1 座標の様式で新たなワークシートに転写する
'''
import os
import sys
import traceback
try:
    import openpyxl
except:
    sys.exit("「pip install openpyxl」を実行して下さい")

def usage(err=""):

    if err:
        print(err, file=sys.stderr)
    msg = '''
 使用方法  python {} [-h] [-o output] input

    -h         ヘルプ
    -o output  出力ファイルの名前。既定は画面（タブ区切り）
    input      入力するワークブックの名前【必須】

 注意点 : 座標データは input の Sheet1 に存在する事''' \
    .format(os.path.basename(__file__))
    sys.exit(msg)

def change_format(src, dst):

    wb_i = openpyxl.load_workbook(src) # 入力用の workbook オブジェクト

    if dst != "":
        type = dst.split(".")[-1]
        if type in ["xls", "xlsx"]: 
            wb_o = openpyxl.Workbook() # 出力用の workbook オブジェクト
            ws = wb_o.worksheets[0]  # ワークシート名は既定の「Sheet」
            r = 0; type = "xls"
        else:
            delim = "," if type == "csv" else "\t"
            f = open(dst, "w")
    else:
        type = "tsv"
        delim = "\t"
        f = sys.stdout

    isHeader = True
    for cols in wb_i["Sheet1"].rows:
        if isHeader:
            kp = cols[0].value # 距離標ないし支川名 (?)
            np = cols[5].value # 座標点数
            if np is None:
                if type == "xls":
                    r += 1
                    ws.cell(r, 1).value = kp # take,tada,zen って支川名 (?)
                else:
                    print(kp, file=f)
                continue # isHeader は True のまま
            count_max = int((np + 3) / 4) # 座標が入力されている行数
            count = 0; xys = []
            isHeader = False
        else:
            for i in range(0, 7, 2): # i = 0, 2, 4, 6
                value_1 = cols[i].value
                if value_1 is None: break
                xys.append([value_1, cols[i + 1].value]) # 二次元配列に追加
            count += 1
            if count == count_max:
                if type == "xls":
                    r += 1
                    ws.cell(r, 1).value = kp
                    ws.cell(r, 2).value = -9999 # 良いのか？
                    for xy in xys:
                        r += 1
                        ws.cell(r, 1).value = xy[0]
                        ws.cell(r, 2).value = xy[1]
                else:
                    print(kp, -9999, sep=delim, file=f)
                    for xy in xys:
                        print(xy[0], xy[1], sep=delim, file=f)
                isHeader = True

    if type == "xls":
        wb_o.save(dst)
        wb_i.close()
        wb_o.close()
    else:
        f.close()

if __name__ == "__main__":

    i = 1; src = ""; dst = ""
    while i < len(sys.argv):
        arg = sys.argv[i]
        if arg == "-h":
            usage()
        if arg == "-o":
            i += 1; dst = sys.argv[i]
        else:
            src = arg
            if src.split(".")[-1] != "xlsx":
                usage(f"ERROR {arg} はワークブックでない")
        i += 1
    if src != "":
        try:
            change_format(src, dst)
        except Exception as e:
            print("\n", traceback.format_exception_only(type(e), e)[0])
    else:
        usage("ERROR  入力ファイルが未指定")