
import os
import sys
import json
import glob
import csv
import openpyxl
import shutil

class CrossSections:

    csObjs = [] # list of CrossSection

    def __init__(self, src):

        type = src["type"]
        if   type == "NK":   NK.  import_from(self, src)
        elif type == "CTI":  CTI. import_from(self, src)
        elif type == "MLIT": MLIT.import_from(self, src)
        elif type == "JSON": JSON.import_from(self, src)
        else: raise Error(f"No such type : '{type}'")

    def export_to(self, dst):

        type = dst["type"]
        if   type == "NK":   NK.  export_to(self, dst)
        elif type == "CTI":  CTI. export_to(self, dst)
        elif type == "MLIT": MLIT.export_to(self, dst)
        elif type == "JSON": JSON.export_to(self, dst)
        else: raise Error(f"No such type : '{type}'")

    @staticmethod
    def to_valiant(s):

        try:    return float(s)
        except: return s

    @staticmethod
    def csv_2_xls(input, output, sheet):

        if os.path.exists(output):
            exist = True
            wb = openpyxl.load_workbook(output)
            ws = wb[sheet]
        else:
            exist = False
            wb = openpyxl.Workbook()
            ws = wb.active

        r = 1
        with open(input) as f:
            for cols in csv.reader(f):
                for i, col in enumerate(cols):
                    ws.cell(r, i + 1).value = CrossSections.to_valiant(col)
                r += 1

        if not exist: ws.title = sheet
        wb.save(output)
        wb.close()

class CrossSection:

    cs = {}

    def setDistance(self, distance_lower):
        interval = self.cs["distance"]
        distance = int((interval + distance_lower) * 1000 + 0.5) / 1000
        self.cs["distance"] = distance
        return distance

class CTI(CrossSections):

    def import_from(self, src):

        input = src["file"]
        if not "." in input: raise Error(f"Invalid path : '{input}'")
        ext = input.split(".")[-1].lower()
        if ext == "xlsx":
            if "sheet" in src: sheet = src["sheet"]
            else: raise Error(f"Sheet name is not given : '{input}'")
            CTI.import_from_xls(self, input, sheet)
        elif ext == "csv":
            CTI.import_from_csv(self, input)
        else:
            CTI.import_from_txt(self, input)

        if "rough" in src:
            rough = src["rough"]
            input = rough["file"]
            if not "." in input: raise Error(f"Invalid path : '{input}'")
            ext = input.split(".")[-1].lower()
            if ext == "xlsx":
                if "sheet" in src: sheet = rough["sheet"]
                else: raise Error(f"Sheet name is not given : '{input}'")
                CTI.import_from_xls_2(self, input, sheet)
            elif ext == "csv":
                CTI.import_from_csv_2(self, input)
            else:
                raise Error(f"Not support CTI.import_from_{ext}_2()")

    def import_from_xls(self, input, sheet):

        wb = openpyxl.load_workbook(input)
        ws = wb[sheet]
        r = 1; distance = 0
        while ws.cell(r, 2).value is not None:
            obj = CTI_()
            r = obj.import_from_xls(ws, r)
            distance = obj.setDistance(distance)
            self.csObjs.append(obj)
        wb.close()

    def import_from_xls_2(self, input, sheet):

        wb = openpyxl.load_workbook(input)
        ws = wb[sheet]
        for i, cs in enumerate(self.csObjs):
            r = i + 2
            name = ws.cell(r, 1).value
            if name == cs.name:
                nhl = ws.cell(r, 2).value # left-side floodplain
                nl  = ws.cell(r, 3).value # lower-channel
                nhr = ws.cell(r, 4).value # right-side floodplain
                ns  = [nhl, nl, nhr]
                if cs.lr != cs.llr: # lower-channel & floodplain is separated
                    cs["roughness"] = {"changeAt": cs.llr, "values": ns}
                else:
                    cs["roughness"] = max(ns)
            else:
                raise Error(f"name mismatch: {name} {cs.name}")
        wb.close()

    def import_from_csv(self, input):

        with open(input) as f:
            distance = 0
            for header in csv.reader(f):
                nmax = header[5] # nos of point
                obj = CTI_()
                CTI_.import_from_csv(obj, f, header, nmax)
                distance = obj.setDistance(distance)
                self.csObjs.append(obj)

    def import_from_csv_2(self, input):

        with open(input) as f:
            reader = csv.reader(f)
            try:
                next(f)
                for obj in self.csObjs:
                    cs = obj.cs
                    cols = next(reader)
                    name = cols[0]
                    if name == cs["name"]:
                        nhl = float(cols[1]) if cols[1] else 0 # left-side floodplain
                        nl  = float(cols[2]) if cols[2] else 0 # lower-channel
                        nhr = float(cols[3]) if cols[3] else 0 # right-side floodplain
                        ns  = [nhl, nl, nhr]
                        llr = cs["lowerChannel"]
                        if llr != cs["trimAt"]:
                            cs["roughness"] = {"changeAt": llr, "values": ns}
                        else:
                            cs["roughness"] = max(ns)
                    else:
                        raise Error(f'name mismatch: {name} {cs["name"]}')
            except:
                raise Error(f"EOF read unexpectedly: {input}")

    def import_from_txt(self, input):

        with open(input) as f:
            distance = 0
            for header in f:
                nmax = int(header[50:60]) # nos of point
                obj = CTI_()
                CTI_.import_from_txt(obj, f, header, nmax)
                distance = obj.setDistance(distance)
                self.csObjs.append(obj)

    def export_to(self, dst):

        output = dst["file"]
        if not "." in output: raise Error(f"Invalid path : '{output}'")
        ext = output.split(".")[-1].lower()
        if ext == "xlsx":
            if "sheet" in dst: sheet = dst["sheet"]
            else: raise Error(f"Sheet name is not given : '{output}'")
            CTI.export_to_xls(self, output, sheet)
        elif ext == "csv":
            CTI.export_to_csv(self, output)
        else:
            CTI.export_to_txt(self, output)

        if "rough" in dst:
            rough = dst["rough"]
            output = rough["file"]
            if not "." in output: raise Error(f"Invalid path : '{output}'")
            ext = output.split(".")[-1].lower()
            if ext == "xlsx":
                if "sheet" in dst: sheet = rough["sheet"]
                else: raise Error(f"Sheet name is not given : '{output}'")
                CTI.export_to_xls_2(self, output, sheet)
            elif ext == "csv":
                CTI.export_to_csv_2(self, output)
            else:
                raise Error(f"Not support CTI.export_to_{ext}_2()")

    def export_to_txt(self, output, csv=False):

        f = open(output, "w")
        x_last = 0
        for csObj in self.csObjs:
            cs = csObj.cs
            x   = cs["distance"]
            lr  = cs["trimAt"]
            llr = cs["lowerChannel"]
            c   = cs["cordinates"]
            n = len(c)
            d = int((x - x_last) * 1000 + 0.5) / 1000 # interval
            if csv:
                slr  = "{}{:5d}".format( lr[0] + 1,  lr[1] + 1) 
                sllr = "{}{:5d}".format(llr[0] + 1, llr[1] + 1)
                print(cs["name"], d, ",,", n, sllr, slr,
                        sep=",", file=f)
            else:
                print("{:10}{:10.3f}{:40d}{:5d}{:5d}{:5d}{:5d}" \
                    .format(cs["name"], d, n, llr[0]+1, llr[1]+1, lr[0]+1, lr[1]+1),
                    file=f)
            b = 0
            for _ in range(int((n + 3) / 4)):
                e = min(b + 4, n)
                for j in range(b, e):
                    if csv:
                        if j == b: s  =  f"{c[j][0]},{c[j][1]}"
                        else:      s += f",{c[j][0]},{c[j][1]}"
                    else:
                        if j == b: s  = f"{c[j][0]:10.3f}{c[j][1]:10.3f}"
                        else:      s += f"{c[j][0]:10.3f}{c[j][1]:10.3f}"
                print(s, file=f)
                b = e
            x_last = x
        f.close()

    def export_to_csv(self, output):

        CTI.export_to_txt(self, output, csv=True)

    def export_to_xls(self, output, sheet):
        
        CTI.export_to_txt(self, "tmpfile.csv", csv=True)
        self.csv_2_xls("tmpfile.csv", output, sheet)
        os.remove("tmpfile.csv")

    def export_to_csv_2(self, output):

        if not "roughness" in self.csObjs[0].cs:
            print(f"File '{output}' not create, \
                because non-exist datas of roughness", file=sys.stderr)
            return False

        with open(output, "w") as f:
            print("距離標,左岸高水敷,低水路,左岸高水敷", file=f)
            for obj in self.csObjs:
                cs = obj.cs
                rough = cs["roughness"]
                if type(rough) is object:
                    if rough["changeAt"] == cs["lowerChannel"]:
                        values = rough["values"]
                        print(f'{cs["name"]},{values[0]},{values[1]},{values[2]}', file=f)
                    else:
                        print(f'{cs["name"]},?,?,?', file=f) # give up
                else:
                    print(f'{cs["name"]},{rough},{rough},{rough}', file=f)
        return True

    def export_to_xls_2(self, output, sheet):

        if CTI.export_to_csv_2(self, "tmpfile.csv"):
            self.csv_2_xls("tmpfile.csv", output, sheet)
            os.remove("tmpfile.csv")

class CTI_(CrossSection):

    def __init__(self):
        pass

    def import_from_xls(self, ws, rb):

        nmax = ws.cell(rb, 6).value
        try:
            llr = ws.cell(rb, 7).value.split(" ") # node nos of lower-channel
            lr  = ws.cell(rb, 8).value.split(" ") # node nos of effective range
        except:
            llr = (1, nmax); lr = (1, nmax)
        name         = str(ws.cell(rb,1).value).strip()
        interval     = ws.cell(rb,2).value
        lowerChannel = (int(llr[0]) - 1, int(llr[-1]) - 1)
        trimAt       = (int(lr[ 0]) - 1, int(lr[ -1]) - 1)

        rb += 1; re = rb + int((nmax + 3) / 4)
        n = 0; cordinates = []
        for r in range(rb, re):
            for c in range(1, 8, 2): # c = 1, 3, 5, 7
                h = ws.cell(r, c    ).value # holizontal
                v = ws.cell(r, c + 1).value # vertical
                cordinates.append((h, v))
                n += 1
                if n == nmax:
                    break
        cordinates = tuple(cordinates)
        self.cs = {
            "name":         name,
            "distance":     interval,
            "lowerChannel": lowerChannel,
            "trimAt":       trimAt,
            "cordinates":   cordinates
        }
        return re

    def import_from_txt(self, f, header, nmax):

        ll = int(header[60:65]) - 1 # node no of lower-channel (left)
        lr = int(header[65:70]) - 1 # node no of lower-channel (right)
        l  = int(header[70:75]) - 1 # node no of effective range (left)
        r  = int(header[75:80]) - 1 # node no of effective range (right)

        name         = header[:10].strip()
        interval     = float(header[10:20])
        lowerChannel = (ll, lr)
        trimAt       = (l,  r )

        n = 0; cordinates = []
        for _ in range(int((nmax + 3) / 4)):
            line = next(f)
            for i in range(0, 80, 20):
                j = i + 10
                k = j + 10
                h = float(line[i:j]) # holizontal
                v = float(line[j:k]) # vertical
                cordinates.append((h, v))
                n += 1
                if n == nmax:
                    break
        cordinates = tuple(cordinates)
        self.cs = {
            "name":         name,
            "distance":     interval,
            "lowerChannel": lowerChannel,
            "trimAt":       trimAt,
            "cordinates":   cordinates
        }

    def import_from_csv(self, f, header, nmax):

        llr = header[6].split(" ")
        lr  = header[7].split(" ")

        name         = header[0].strip()
        interval     = float(header[1])
        lowerChannel = (llr[0] - 1, llr[-1] - 1)
        trimAt       = ( lr[0] - 1,  lr[-1] - 1)

        n = 0; cordinates = []
        for _ in range(int((nmax + 3) / 4)):
            line = csv.reader(f)
            for i in range(0, 8, 2): # 0, 3, 5, 7
                j = i + 1
                h = float(line[i]) # holizontal
                v = float(line[j]) # vertical
                cordinates.append((h, v))
                n += 1
                if n == nmax:
                    break
        cordinates = tuple(cordinates)
        self.cs = {
            "name":         name,
            "distance":     interval,
            "lowerChannel": lowerChannel,
            "trimAt":       trimAt,
            "cordinates":   cordinates
        }

class NK(CrossSections):

    def import_from(self, src):

        input = src["file"]
        if not "." in input: raise Error(f"Invalid path : '{input}'")
        ext = input.split(".")[-1].lower()
        if ext == "xlsm":
            NK.import_from_xls(self, input)
        else:
            raise Error(f"Class NK not has method import_from_{ext}() : {input}")

    def import_from_xls(self, input):

        wb = openpyxl.load_workbook(input)
        ws = wb["横断データ"]
        r = 2
        while ws.cell(r, 2).value is not None:
            obj = NK_()
            r = obj.import_from_xls(ws, r)
            self.csObjs.append(obj)

        ws = wb["縦断データ"]
        r = 2; idx = 0; distance = 0
        while ws.cell(r, 2).value is not None:
            distance += ws.cell(r, 3).value
            self.csObjs[idx].cs["distance"] = distance
            r += 1; idx += 1

        wb.close()

    def export_to(self, dst):

        output = dst["file"]
        if not "." in output: raise Error(f"Invalid path : '{output}'")
        ext = output.split(".")[-1].lower()
        if ext == "xlsm":
            NK.export_to_xls(self, output)
        else:
            raise Error(f"Class NK not has method export_to_{ext}() : {output}")

    def export_to_xls(self, output):

        shutil.copyfile("template_Q2DFNU.xlsm", output)
        wb = openpyxl.load_workbook(output, keep_vba=True)
        ws = wb["横断データ"]
        ws.cell(1, 1).value = "lr"
        ws.cell(1, 2).value = "L"
        ws.cell(1, 3).value = "Z"
        ws.cell(1, 4).value = "粗度係数"
        ws.cell(1, 5).value = "境界混合係数"
        ws.cell(1, 6).value = "樹高"
        ws.cell(1, 7).value = "樹木境界混合係数"
        if "roughness" in self.csObjs[0].cs: hasRough = True
        else:                                hasRough = False
        r = 1
        for csObj in self.csObjs:
            cs = csObj.cs
            llr = cs["lowerChannel"]
            lr  = cs["trimAt"]
            cordinates = cs["cordinates"]
            if hasRough:
                rough = cs["roughness"]
                if type(rough) is object:
                    ats = rough["changeAt"] + [lr[1]]
                    ns  = rough["values"]
                else:
                    ats = [lr[1]]
                    ns  = [rough]
            else:   ats = [len(cordinates)] # never match
            r += 1; j = 0
            ws.cell(r, 2).value = self.to_valiant(cs["name"])
            ws.cell(r, 3).value = len(cordinates)
            ws.cell(r, 8).value = "*"
            for i, hv in enumerate(cordinates):
                if   i == lr[ 0]: col_1 = "l"
                elif i == lr[ 1]: col_1 = "r"
                elif i == llr[0]: col_1 = "lm"
                elif i == llr[1]: col_1 = "rm"
                else:             col_1 = ""
                r += 1
                ws.cell(r, 1).value = col_1
                ws.cell(r, 2).value = hv[0]
                ws.cell(r, 3).value = hv[1]
                if i == ats[j]:
                    ws.cell(r, 4).value = ns[j]
                    j += 1

        ws = wb["縦断データ"]
        ws.cell(1, 1).value = "断面番号"
        ws.cell(1, 2).value = "断面"
        ws.cell(1, 3).value = "区間距離(m)"
        ws.cell(1, 4).value = "流量(m3/s)"
        r = 1; xLast = 0
        for i, csObj in enumerate(self.csObjs):
            cs = csObj.cs
            distance = cs["distance"]
            r += 1
            ws.cell(r, 1).value = i + 1
            ws.cell(r, 2).value = self.to_valiant(cs["name"])
            ws.cell(r, 3).value = distance - xLast
            xLast = distance

        wb.save(output)
        wb.close()

class NK_(CrossSection):

    def __init__(self):
        pass

    def import_from_xls(self, ws, r):

        name = str(ws.cell(r, 2).value)
        np = ws.cell(r, 3).value
        lr = [0, np - 1]; llr = [None, None]; hvs = []
        ns = []; ats = []
        for i in range(np):
            r += 1
            col_1 = ws.cell(r, 1).value
            if   col_1 == "l":  lr[ 0] = i
            elif col_1 == "r":  lr[ 1] = i
            elif col_1 == "lm": llr[1] = i
            elif col_1 == "rm": llr[1] = i
            h = ws.cell(r, 2).value # holizontal
            v = ws.cell(r, 3).value # vertical
            n = ws.cell(r, 4).value # roughness coef.
            if n is not None:
                ns.append(n); ats.append(i)
            hvs.append([h, v])
        if llr[0] is None: llr[0] = lr[0]
        if llr[1] is None: llr[1] = lr[1]
        self.cs = {
            "name":         name,
            "lowerChannel": tuple(llr),
            "trimAt":       tuple(lr),
            "cordinates":   tuple(hvs)
        } # property "distance" is set with other sheet

        sz = len(ns)
        if sz == 1:
            self.cs["roughness"] = ns[0]
        elif sz != 0:
            self.cs["roughness"] = {"changeAt": ns[:-1], "values": ns}

        return r + 1

class MLIT(CrossSections):

    def import_from(self, src):

        input = src["file"]
        if os.path.isfile(input): # concatenate WZA4[0-9]+.csv
            with open(input) as f:
                while True:
                    try:
                        obj = MLIT_(f)
                        self.csObjs.append(obj)
                        if obj.hasStructure:
                            next(f)
                    except:
                        break
        else: # input/WZA4[0-9]+.csv
            for file in glob.glob(input + "/WZAA4*.CSV"):
                with open(file) as f:
                    obj = MLIT_(f)
                    self.csObjs.append(obj)

    def export_to(self, dst):

        output = dst["file"]
        distance = 0
        if MLIT.isFile(output):
            with open(output, "w") as f:
                for obj in self.csObjs:
                    distance = MLIT_.export_to(obj, f, distance)
        else:
            for i, obj in enumerate(self.csObjs):
                with open(output + "/WZAA4{:03d}.CSV".format(i + 1), "w") as f:
                    distance = MLIT_.export_to(obj, f, distance)

    @staticmethod
    def isFile(path):
        if "." in path:
            ext = path.split(".")[-1].lower()
            return True if ext == "csv" else False
        else:
            False

class MLIT_(CrossSection):

    hasStructure = False

    def __init__(self, f):
        ''' Create New MLIT_ Object
        Arg
            f (obj): file handle
        '''
        reader = csv.reader(f)
        cols = next(reader) 
        name = cols[0].strip()
        interval = float(cols[1])
        self.hasStructure = True if int(cols[11]) > 0 else False
        ts = []; hvs = []; hs = []
        for _ in range(int(cols[6])):
            vals = next(reader)
            h = float(vals[1]) # horizontal
            v = float(vals[2]) # vertical
            ts.append(int(vals[0]))
            hvs.append((h, v))
            hs.append(h)
        sz  = len(hvs)
        lr  = (ts.index( 1) if  1 in ts else 0,
               ts.index(18) if 18 in ts else sz - 1)
        llr = self.set_llr(ts, hs, sz, lr)

        self.cs = {
            "name":         name,
            "distance":     interval,
            "lowerChannel": llr,
            "trimAt":       lr,
            "cordinates":   tuple(hvs)
        }

    @staticmethod
    def set_llr(ts, hs, sz, lr):

        # ts is series of node type
        #   12 : stakes at the water's edge
        #   13 : border between lower-channel and flood-plain

        ts_r = list(reversed(ts))
        c12 = ts.count(12); c13 = ts.count(13)
        if   c13 == 2:
            return (ts.index(13), sz - ts_r.index(13))
        elif c13 == 1:
            i13 = ts.index(13); h13 = hs[i13]
            if c12 == 2:
                i12s = (ts.index(12), sz - ts_r.index(12))
                d12s = [abs(hs[i] - h13) for i in i12s]
                if d12s[0] < d12s[1]: return (i13, i12s[1])
                else:                 return (i12s[0], i13)
        if   c12 == 2:
            return (ts.index(12), sz - ts_r.index(12))
        elif c12 == 1:
            i12 = ts.index(12); h12 = hs[i12]
            d12s = (abs(h12 - lr[0]), abs(h12 - lr[1]))
            if d12s[0] < d12s[1]: return (i12,    lr[1])
            else:                 return (lr[0], i12   )
        return lr

    def export_to(self, f, distance_last):

        cs = self.cs
        name       = cs["name"]
        distance   = cs["distance"]
        llr        = cs["lowerChannel"]
        lr         = cs["trimAt"]
        cordinates = cs["cordinates"]
        interval = distance - distance_last
        s = ",,,,,,0,,,ATTENSION_THAT_STRUCTURE_FLAG_SET_TO_ZERO"
        print(name, interval, s, sep=",", file=f)
        for i, cordinate in enumerate(cordinates):
            if   i == lr[0]:  t =  1 # effective range (left  side)
            elif i == llr[1]: t = 18 # effective range (right side) 
            elif i in lr    : t = 13 # boundar between lower-channel and floodplain
            else:             t =  0
            print(t, cordinate[0], cordinate[1], sep=",", file=f)
        return distance

class JSON(CrossSections):

    def import_from(self, src):

        with open(src["file"]) as f:
            js = json.load(f)
        for j in js:
            obj = JSON_()
            obj.cs = j
            self.csObjs.append(obj)

    def export_to(self, dst):

        objs = []
        for obj in self.csObjs:
            objs.append(obj.cs)
        with open(dst["file"], "w") as  f:
            json.dump(objs, f)

class JSON_(CrossSection):

    def __init__(self):
        pass

class Error(Exception):
    pass